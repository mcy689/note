import easyocr
import re
import os
import time
from openpyxl import Workbook

reader = easyocr.Reader(['ch_sim', 'en'])


def readeOcr(fileName):
    result = reader.readtext(fileName)
    fullStr = ""
    for x in result:
        for everyOne in x:
            if isinstance(everyOne, str):
                result = everyOne.replace(' ', '')
                result = result.replace('\n', '')
                fullStr += result+"FFF"
    result = re.findall(
        r'FFF(DPK[0-9]+)FFF([\u4e00-\u9fa5]+)[0-9]{0,1}FFF(1[3456789]\d{9})FFF', fullStr)
    if len(result) == 0:
        result = re.findall(
            r'FFF(SF[0-9]+)FFF.*FFF([\u4e00-\u9fa5]+)FFF([0-9]{3}\*{4}[0-9]{4})FFF', fullStr)
    if len(result) == 0:
        return ()
    return result[0]

# 读取当前运行环境下的 images 文件夹下所有图片
path = os.getcwd()+os.sep+"images"+os.sep
if not os.path.exists(path):
    os.mkdir(path)
nowPath = path+time.strftime('%Y-%m-%d', time.localtime(time.time()))+os.sep
if not os.path.exists(nowPath):
    os.mkdir(nowPath)

files = os.listdir(path)
dataResult = []
imgeExt = (".jpg", ".png", ".jpeg")
for file in files:
    filePath = path+file
    if not os.path.isdir(filePath):
        suffixExt = os.path.splitext(filePath)[-1]
        if suffixExt in imgeExt:
            result = readeOcr(filePath)
            if len(result) == 0:
                continue
            dataResult.append(result)
            fileName = time.strftime('%Y%m%d', time.localtime(time.time()))
            os.rename(filePath, nowPath+fileName+"-"+result[1]+suffixExt)

wb = Workbook()
ws = wb.active
ws['A1'] = "物流单号"
ws['B1'] = "姓名"
ws['C1'] = "手机号"
for data in dataResult:
    ws.append(data)

wb.save(path+"识别成功物流信息.xlsx")
